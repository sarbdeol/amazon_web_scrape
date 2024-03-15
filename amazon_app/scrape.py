import pandas as pd
import time
import requests
from bs4 import BeautifulSoup
import json
import csv
import os
from urllib.parse import quote
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from .models import Header
from .models import ScrapeDataCount
header_instance = Header.objects.first()  # Assuming there's only one instance
# print(header_instance)
# Define all headers
headers = [
    'store_name', 'Asin', 'Category', 'Title', 'MRP', 'Price', 'ShippingPrice',
    'Manufacturer', 'StarRating', 'IsPrime', 'CustomerReview', 'Weight', 'Color',
    'Size', 'PackageDimensions', 'ItemModelNumber', 'Department', 'DateFirstAvailable',
    'BestSellersRank', 'BulletPoints', 'ImageURLs', 'ProductUrl', 'VedioLinks', 'DeliveryTime', 'Description'
]

# Create a dictionary to store the enabled status of each header
header_status = {header: getattr(header_instance, header.lower(), False) for header in headers}
#* * * * * /usr/bin/python3 /root/Scraper/Downloads/scrape3.py >> /root/Scraper/Downloads/cron3.log 3>&1
def add_data_to_excel(data):
    # Open the workbook or create a new one if it doesn't exist
    try:
        wb = openpyxl.load_workbook('output_asin3.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Select the active worksheet (the first sheet by default)
    ws = wb.active

    # Append the data to the next available row
    next_row = ws.max_row + 1
    for col, (key, value) in enumerate(data.items(), start=1):
        ws.cell(row=next_row, column=col, value=value)

    # Save the workbook
    wb.save('output_asin3.xlsx')





def add_data_to_csv(data):
    try:
        # Open the CSV file in append mode, create a new one if it doesn't exist
        with open('download/output.csv', 'a', newline='') as csvfile:
            # Create a CSV writer object
            writer = csv.writer(csvfile)
            
            # Write data to the CSV file
            writer.writerow(data.values())

            # Increment the scrape data count
            scrape_data_count = ScrapeDataCount.objects.first()
            if scrape_data_count:
                scrape_data_count.count += 1
                scrape_data_count.save()
            else:
                ScrapeDataCount.objects.create(count=1)

        print("Data added successfully.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
# Get the current working directory
current_directory = os.getcwd()

# API token for scrape.do
token = "2594e5d442b4457aa0e37b651200277726751d1e518"

# Function to get URL with token
def get_url(url):
    targetUrl = quote(url)
    proxy_url = "http://api.scrape.do?token={}&url={}&geoCode=in".format(token, targetUrl)
    return proxy_url
# Function to append data to an Excel file
data = {}
count=0
def scrape_amazon(asin_value):
    global count
    count+=1
    print(f'asin {count}:',asin_value)
    product_url = f'https://www.amazon.in/dp/{asin_value}/'
    proxy_url = get_url(product_url)
    print(proxy_url)
    response = requests.get(proxy_url)
    if response.status_code == 200:
        # Parse HTML content
        # Assuming `response` contains the HTML content
        data['Asin'] = asin_value
        soup = BeautifulSoup(response.content, 'html.parser')
        if header_status.get('store_name'):
            try:
                # Find the span containing the text 'Sold by'
                StoreName = soup.find('a', id='sellerProfileTriggerId').get_text(strip=True)
                data['StoreName'] = StoreName
            except :
                # print("Error occurred:", e)
                data['StoreName'] = ''
    
        if header_status.get('Category'):
            try:
                Categories = soup.find_all('a', class_='a-link-normal a-color-tertiary')
                # Extract category texts and join them with '>'
                category = ' > '.join(category.get_text(strip=True) for category in Categories)    
                data['Category'] = category
            except:
                data['Category'] =  ''
        # print('category',category)
        if header_status.get('Title'):
            try:
                title_element = soup.find('span', id='productTitle')
                title = title_element.get_text(strip=True).strip()
                data['Title']=title
            except:
                data['Title']= ''
        # print('title',title)
        if header_status.get('MRP'):
            try:
                mrp = soup.find('span', class_='a-price a-text-price').find('span').get_text().strip()
                data['MRP']=mrp.replace('₹','')
            except:
                data['MRP']=''
        if header_status.get('ShippingPrice'):
            try:
                ShippingPrice = soup.select_one('#mir-layout-DELIVERY_BLOCK-slot-PRIMARY_DELIVERY_MESSAGE_LARGE .a-link-normal:nth-child(1)').get_text().strip()
                data['ShippingPrice']=ShippingPrice
            except:
                data['ShippingPrice'] = ''
        # print('ShippingPrice',ShippingPrice)
                
        if header_status.get('Manufacturer'):
            try:
                # Try to find the table row containing the manufacturer's information
                manufacturer_row = soup.find('th', string=' Manufacturer ').find_parent('tr')
                if manufacturer_row:
                    # If the manufacturer row is found, extract the manufacturer's name from the table row
                    Manufacturer = manufacturer_row.find('td').get_text(strip=True)
                # If the manufacturer row is not found
                if manufacturer_row is None:
                    # Find all spans with class 'a-text-bold'
                    bold_spans = soup.find_all('span', class_='a-text-bold')
                    # Search for the span containing the text 'Manufacturer'
                    for span in bold_spans:
                        if 'Manufacturer' in span.get_text():
                            # Get the next sibling span which contains the manufacturer's name
                            Manufacturer = span.find_next_sibling('span').get_text(strip=True)
                            break  # Stop searching once the manufacturer's name is found
                # Print the manufacturer's name
                Manufacturer= Manufacturer
                data['Manufacturer']=Manufacturer
            except:
                # Handle the case where an attribute error occurs
                data['Manufacturer'] = ''
        if header_status.get('StarRating'):
            try:
                rating_element = soup.find('span', class_='reviewCountTextLinkedHistogram')
                StarRating = rating_element['title'].split('out')[0]
                data['StarRating']=StarRating
                
            except:
                data['StarRating']=''
        # print('StarRating',StarRating)
        if header_status.get('IsPrime'):
            data['IsPrime']=0
        if header_status.get('CustomerReview'):
            try:
                reviews_element = soup.find('span', id='acrCustomerReviewText')
                CustomerReview = reviews_element.get_text(strip=True)
                data['CustomerReview']=CustomerReview
            except:

                data['CustomerReview'] = ''
        # print('CustomerReview',CustomerReview)
        if header_status.get('Weight'):
            try:
                weight_span = soup.find('th', string=' Item Weight ').find_parent('tr')
                if weight_span:
                    
                    # Get the next sibling span which contains the weight value
                    Weight = weight_span.find('td').get_text(strip=True)
                
                if weight_span is None:
                    bold_spans = soup.find_all('span', class_='a-text-bold')

                    # Search for the span containing the text 'Date First Available'
                    for span in bold_spans:
                        if 'Weight' in span.get_text():
                            # Get the next sibling span which contains the date
                            Weight = span.find_next_sibling('span').get_text(strip=True)
                            # print("Item Weight", Weight)
                            break  # Stop searching once the date is found
                
                Weight=Weight
                data['Weight']=Weight
            except:
            
                data['Weight']=''
        if header_status.get('Color'):
            try:
                Color = soup.find('span', string=' Color : ')

                # Get the next sibling span which contains the weight value
                Color = Color.find_next_sibling('span').get_text(strip=True)
                data['Color']=Color
            except:
                data['Color'] = ''
        # print('Color',Color)
        if header_status.get('Size'):
            data['Size']=''
        if header_status.get('ItemModelNumber'):
            try:
                try:
                    ItemModelNumber_span = soup.find('th', string=' Item part number ').find_parent('tr')
                except:
                    ItemModelNumber_span=None
                if ItemModelNumber_span:


                    # Get the next sibling span which contains the weight value
                    ItemModelNumber = ItemModelNumber_span.find('td').get_text(strip=True).get_text(strip=True)
                if ItemModelNumber_span is None:
                    bold_spans = soup.find_all('span', class_='a-text-bold')

                    # Search for the span containing the text 'Date First Available'
                    for span in bold_spans:
                        if 'Item part number' in span.get_text():
                            # Get the next sibling span which contains the date
                            ItemModelNumber = span.find_next_sibling('span').get_text(strip=True)
                            print("ItemModelNumber", ItemModelNumber)
                            break  # Stop searching once the date is found
                
                ItemModelNumber=ItemModelNumber
                data['ItemModelNumber']=ItemModelNumber
            except:
            
                data['ItemModelNumber']=''
                pass
        if header_status.get('Department'):
            data['Department']=''
        if header_status.get('DateFirstAvailable'):
            try:
                DateFirstAvailable = soup.find('th', string=' Date First Available ').find_parent('tr')

        
                # print('DateFirstAvailable',DateFirstAvailable)
                DateFirstAvailable=DateFirstAvailable.find('td').get_text(strip=True)
            except:
                # Find all spans with class 'a-text-bold'
                bold_spans = soup.find_all('span', class_='a-text-bold')
                # Search for the span containing the text 'Date First Available'
                for span in bold_spans:
                    if 'Date First Available' in span.get_text():
                        # Get the next sibling span which contains the date
                        DateFirstAvailable = span.find_next_sibling('span').get_text(strip=True)
                        print("Date First Available:", DateFirstAvailable)
                        break  # Stop searching once the date is found
                else:
                    DateFirstAvailable=''
                    # If the span containing the text 'Date First Available' is not found
                    # print("Date First Available: Not found")
        if header_status.get('PackageDimensions'):
            try:
                
                try:
                    product_dimensions_row = soup.find('th', string='Product Dimensions').find_parent('tr')

                except:

                    product_dimensions_row=None
                # print('product_dimensions_row',product_dimensions_row)
                if product_dimensions_row is not None:
                    # Get the text from the next sibling td element which contains the product dimensions
                    Dimensions = product_dimensions_row.find('td').get_text(strip=True)
                
                elif product_dimensions_row is None:
                    bold_spans = soup.find_all('span', class_='a-text-bold')

                    # Search for the span containing the text 'Date First Available'
                    for span in bold_spans:
                        if 'Dimensions' in span.get_text():
                            # Get the next sibling span which contains the date
                            Dimensions = span.find_next_sibling('span').get_text(strip=True).split(';')[0]
                            print("Dimensions", Dimensions)
                            break  # Stop searching once the date is found
            
                Dimensions=Dimensions
                data['Dimensions']=Dimensions
            except:
                data['Dimensions'] = ''
        # print('Dimensions',Dimensions)
        if header_status.get('BestSellersRank'):
            try:

                bestsellerrank = soup.find('th', string=' Best Sellers Rank ').find_parent('tr')
                # print('bestsellerrank',bestsellerrank)
                if bestsellerrank is not None:
                    # Find all the next sibling span elements
                    # Find the row containing the Best Sellers Rank information
                    best_sellers_rank_row = soup.find('th', string=' Best Sellers Rank ').find_parent('tr')

                    # Extract the Best Sellers Rank text from the sibling td element
                    best_sellers_rank_text = best_sellers_rank_row.find('td').get_text(strip=True)

                    # Split the Best Sellers Rank text into individual ranks
                    sellerrank_texts = [rank.strip() for rank in best_sellers_rank_text.split('#') if rank.strip()]

                    # Assign the texts to individual variables sellerrank1, sellerrank2, sellerrank3, etc.
                    sellerrank1, sellerrank2, sellerrank3 = sellerrank_texts + [''] * (3 - len(sellerrank_texts))
                if bestsellerrank is None:
                    bold_spans = soup.find_all('span', class_='a-text-bold')

                    # Search for the span containing the text 'Date First Available'
                    for span in bold_spans:
                        if 'Sellers' in span.get_text():
                            # Get the next sibling span which contains the date
                            sellerrank_spans = span.find_next_siblings('span').get_text(strip=True)
                            sellerrank_texts = [span.get_text(strip=True) for span in sellerrank_spans[:4]]
                            print("sellerrank_texts", sellerrank_texts)
                            # Assign the texts to individual variables sellerrank1, sellerrank2, sellerrank3, sellerrank4
                            sellerrank1, sellerrank2, sellerrank3 = sellerrank_texts + [''] * (3 - len(sellerrank_texts))
                            break  # Stop searching once the date is found
                
                sellerrank1=sellerrank1
                data['BestSellersRank1'] = sellerrank1
                data['BestSellersRank2'] = sellerrank2
                data['BestSellersRank3'] = sellerrank3
            except:
                data['BestSellersRank1']=''
                data['BestSellersRank2']=''
                data['BestSellersRank3']= ''

        # print('sellerrank1, sellerrank2, sellerrank3',sellerrank1, sellerrank2, sellerrank3)
        if header_status.get('BulletPoints'):
            try:
                # Find the div containing bullet points
                bullets_div = soup.find('div', id='feature-bullets')

                # Extract bullet point texts from the span elements
                bullet_spans = bullets_div.find_all('span', class_='a-list-item')
                bullet_texts = [span.get_text(strip=True) for span in bullet_spans]

                # Fill in empty strings if less than 10 bullet points
                while len(bullet_texts) < 10:
                    bullet_texts.append('')

                # Slice the list to get the first 10 bullet points
                bullet_texts = bullet_texts[:10]

                # Assign bullet points to individual variables
                bulletpoint1, bulletpoint2, bulletpoint3, bulletpoint4, bulletpoint5, bulletpoint6, bulletpoint7, bulletpoint8, bulletpoint9, bulletpoint10 = bullet_texts
                if bullet_texts:
                    for i, bullet_text in enumerate(bullet_texts, start=1):
                        data[f'Bulletpoint{i}'] = bullet_text
            except:
                bulletpoint1 = bulletpoint2 = bulletpoint3 = bulletpoint4 = bulletpoint5 = bulletpoint6 = bulletpoint7 = bulletpoint8 = bulletpoint9 = bulletpoint10 = ''
                
                for i, bullet_text in enumerate(bullet_texts, start=1):
                    data[f'Bulletpoint{i}'] = bullet_text
        # print('bulletpoint done')
        if header_status.get('ImageURLs'):
            try:
                # Find the img tag
                img_tag = soup.find('div',id='imgTagWrapperId').find('img')

                # Extract the value of the data-a-dynamic-image attribute
                data_attribute = img_tag.get('data-a-dynamic-image')
                # print(data_attribute)
            
                # Convert the data attribute string to a dictionary
                image_data = json.loads(data_attribute)

                # Extract image URLs from the dictionary
                image_urls = list(image_data.keys())

                # Fill in empty strings if less than 10 image URLs
                while len(image_urls) < 10:
                    image_urls.append('')

                # Slice the list to ensure it has exactly 10 elements
                image_urls = image_urls[:10]
                # print(image_urls)
                # Assign image URLs to individual variables
                if image_urls:
                    for i, image_url in enumerate(image_urls, start=1):
                        data[f'ImageURL{i}'] = image_url
            
            except:
                image_element = soup.find('div', id='imgTagWrapperId')
                if image_element:
                    data['ImageURL1'] = image_element.find('img')['src']
                else:
                    data['ImageURL1'] = ''

                # Set ImageURL2 to ImageURL10 to empty strings
                for i in range(2, 11):
                    data[f'ImageURL{i}'] = ''
        
           
        
        if header_status.get('ProductUrl'):
            # Find the canonical link tag
            canonical_link = soup.find('link', rel='canonical')

            # Extract the href attribute value if the canonical link tag exists
            if canonical_link:
                ProductUrl = canonical_link.get('href')
                data['ProductUrl']=ProductUrl
            else:
                data['ProductUrl']=product_url
            # print('ProductUrl',ProductUrl)
        

        if header_status.get('DeliveryTime'):
            DeliveryTime = soup.select_one('#mir-layout-DELIVERY_BLOCK-slot-PRIMARY_DELIVERY_MESSAGE_LARGE .a-text-bold')

            # Extract the text content
            if DeliveryTime:
                delivery_date = DeliveryTime.get_text(strip=True)
                data['DeliveryTime']=delivery_date
                # print("Delivery Date:", delivery_date)
            else:
                data['DeliveryTime']=''
                # print("Delivery Date: not found.")
        if header_status.get('Description'):
            try:
                description_element = soup.find('div', id='productDescription')
                description = description_element.get_text()
                data['Description']=description
            except:
                description_meta = soup.find('meta', attrs={'name': 'description'})

                # Extract the content attribute value if the meta tag exists
                if description_meta:
                    description = description_meta.get('content')
                    # print("Description:", description)
                else:
                    description=''
                data['Description'] =description  
        if header_status.get('Price'):     
            try:
                price_whole_span = soup.find('span', class_='a-price-whole')
                price_fraction_span = soup.find('span', class_='a-price-fraction')
                # Get the text from the price_whole_span and price_fraction_span
                price = price_whole_span.get_text(strip=True).replace('.', '') + '.' + price_fraction_span.get_text(strip=True)
                data['Price']=price

            except:
                data['Price']=''
            return data
            
            
def scrape_asin_batch(asin_values):
    for asin_value in asin_values:
        if asin_value:
            data=scrape_amazon(asin_value)
            try:
                add_data_to_csv(data)
                print(f'added {asin_value}')
            except Exception as e:
                print(f'adding error {asin_value}:',e)
                break
          
            
                

def main(file_path, batch_size):
    print(file_path)
    # Read ASIN values from the Excel file
    try:
        df = pd.read_excel(f'{file_path}', sheet_name='Sheet1', header=None)
    except:
        df = pd.read_excel(f'{file_path}', sheet_name='Sheet', header=None)
    asin_values = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
    print('Reading ASINs')
    print(len(asin_values))

    # Divide ASINs into batches for concurrent scraping
    with ThreadPoolExecutor(max_workers=batch_size) as executor:
        # Submit scraping tasks for each batch of ASINs
        futures = []
        for i in range(0, len(asin_values), batch_size):
            batch = asin_values[i:i+batch_size]
            futures.append(executor.submit(scrape_asin_batch, batch))
            time.sleep(2)  # Adding a delay between submitting batches

        # Wait for all tasks to complete
        for future in as_completed(futures):
            try:
                result = future.result()
                # Process result if needed
            except Exception as e:
                print(f"Error: {e}")

# if __name__ == "__main__":
#     batch_size = 7  # Number of ASINs to scrape concurrently
#     try:
#         main(batch_size)
#     except Exception as e:
#         print(e)
#     print('complete')
